"""
IT Service Desk ETL Pipeline
Portfolio Project - IT Analyst

This script:
1. Extracts raw ticket data from CSV
2. Transforms and cleans the data
3. Loads aggregated summaries into Excel for Power BI / Tableau
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────
# EXTRACT
# ─────────────────────────────────────────
def extract(filepath):
    df = pd.read_csv(filepath, parse_dates=["created_date", "resolved_date"])
    print(f"[EXTRACT] Loaded {len(df)} rows from {filepath}")
    return df


# ─────────────────────────────────────────
# TRANSFORM
# ─────────────────────────────────────────
def transform(df):
    # Drop rows with missing ticket IDs
    df = df.dropna(subset=["ticket_id"])

    # Normalize text fields
    df["category"] = df["category"].str.strip().str.title()
    df["priority"] = df["priority"].str.strip().str.title()
    df["status"] = df["status"].str.strip().str.title()

    # Flag SLA breaches numerically for aggregation
    df["sla_met_flag"] = df["sla_met"].map({"Yes": 1, "No": 0, "Pending": None})

    # Resolved tickets only
    resolved = df[df["status"].isin(["Resolved", "Closed"])].copy()

    # 1. Monthly ticket volume
    monthly = (
        df.groupby(["created_month_num", "created_month"])
        .agg(total_tickets=("ticket_id", "count"))
        .reset_index()
        .sort_values("created_month_num")
    )

    # 2. Category breakdown
    category = (
        df.groupby("category")
        .agg(
            total=("ticket_id", "count"),
            avg_resolution_hrs=("resolution_hours", "mean"),
            sla_compliance_pct=("sla_met_flag", "mean"),
        )
        .reset_index()
        .sort_values("total", ascending=False)
    )
    category["avg_resolution_hrs"] = category["avg_resolution_hrs"].round(1)
    category["sla_compliance_pct"] = (category["sla_compliance_pct"] * 100).round(1)

    # 3. Priority distribution
    priority = (
        df.groupby("priority")
        .agg(
            total=("ticket_id", "count"),
            avg_resolution_hrs=("resolution_hours", "mean"),
            sla_compliance_pct=("sla_met_flag", "mean"),
        )
        .reset_index()
    )
    priority["avg_resolution_hrs"] = priority["avg_resolution_hrs"].round(1)
    priority["sla_compliance_pct"] = (priority["sla_compliance_pct"] * 100).round(1)

    # 4. Analyst performance
    analyst = (
        resolved.groupby("assigned_analyst")
        .agg(
            tickets_resolved=("ticket_id", "count"),
            avg_resolution_hrs=("resolution_hours", "mean"),
            avg_satisfaction=("satisfaction_score", "mean"),
            sla_compliance_pct=("sla_met_flag", "mean"),
        )
        .reset_index()
        .sort_values("tickets_resolved", ascending=False)
    )
    analyst["avg_resolution_hrs"] = analyst["avg_resolution_hrs"].round(1)
    analyst["avg_satisfaction"] = analyst["avg_satisfaction"].round(2)
    analyst["sla_compliance_pct"] = (analyst["sla_compliance_pct"] * 100).round(1)

    # 5. Department volume
    dept = (
        df.groupby("department")
        .agg(
            total_tickets=("ticket_id", "count"),
            avg_satisfaction=("satisfaction_score", "mean"),
        )
        .reset_index()
        .sort_values("total_tickets", ascending=False)
    )
    dept["avg_satisfaction"] = dept["avg_satisfaction"].round(2)

    print("[TRANSFORM] Aggregations complete")
    return df, monthly, category, priority, analyst, dept


# ─────────────────────────────────────────
# LOAD - Write to Excel
# ─────────────────────────────────────────
def style_header(ws, row, num_cols, fill_color="1F4E79", font_color="FFFFFF"):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=font_color, name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="AAAAAA")
        cell.border = Border(bottom=thin)


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def write_sheet(wb, title, df, fill_color="1F4E79"):
    ws = wb.create_sheet(title)
    ws.row_dimensions[1].height = 22

    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())

    style_header(ws, 1, len(df.columns), fill_color)

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else "")
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="EBF3FB")

    auto_width(ws)
    return ws


def load(raw_df, monthly, category, priority, analyst, dept, output_path):
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    write_sheet(wb, "Raw Data", raw_df.drop(columns=["created_month_num"]), fill_color="1F4E79")
    write_sheet(wb, "Monthly Volume", monthly.drop(columns=["created_month_num"]), fill_color="1F6B3C")
    write_sheet(wb, "By Category", category, fill_color="7B2D8B")
    write_sheet(wb, "By Priority", priority, fill_color="B8420A")
    write_sheet(wb, "Analyst Performance", analyst, fill_color="1F4E79")
    write_sheet(wb, "By Department", dept, fill_color="1F6B3C")

    wb.save(output_path)
    print(f"[LOAD] Saved to {output_path}")


# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────
if __name__ == "__main__":
    raw = extract("it_helpdesk_raw.csv")
    raw_df, monthly, category, priority, analyst, dept = transform(raw)
    load(raw_df, monthly, category, priority, analyst, dept, "it_helpdesk_analytics.xlsx")
    print("\n✅ ETL pipeline complete. Open it_helpdesk_analytics.xlsx in Power BI or Tableau.")
